"""
build_workbook.py
------------------
Builds Supply_Chain_KPI_Report.xlsx from the analysis CSVs produced by
analysis/build_summary.py. Raw analysis output is loaded onto "Data_*"
sheets; the "Dashboard" sheet computes every headline figure with live
formulas (SUMIFS / AVERAGEIFS / INDEX-MATCH) referencing those sheets,
plus native Excel charts.

Usage:
    python excel/build_workbook.py
Then recalc (mandatory, per xlsx skill):
    python /mnt/skills/public/xlsx/scripts/recalc.py excel/Supply_Chain_KPI_Report.xlsx
"""
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ANALYSIS_DIR = os.path.join(BASE, "analysis")
OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Supply_Chain_KPI_Report.xlsx")

FONT_NAME = "Arial"
NAVY = "0B1F3A"
TEAL = "1F8A82"
AMBER = "C9791A"
RED = "B23A3A"
LIGHT = "F2F4F7"
GRID = "D9DEE6"

header_font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor=NAVY)
title_font = Font(name=FONT_NAME, size=16, bold=True, color=NAVY)
subtitle_font = Font(name=FONT_NAME, size=10, italic=True, color="666666")
kpi_label_font = Font(name=FONT_NAME, size=9, bold=True, color="666666")
kpi_value_font = Font(name=FONT_NAME, size=20, bold=True, color=NAVY)
thin = Side(style="thin", color=GRID)
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = box


def load_data_sheet(wb, sheet_name, csv_path):
    df = pd.read_csv(csv_path)
    ws = wb.create_sheet(sheet_name)
    ws.append(list(df.columns))
    style_header_row(ws, 1, len(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))
    for c, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)])
        ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 2, 10), 32)
    for r in range(2, len(df) + 2):
        for c in range(1, len(df.columns) + 1):
            ws.cell(row=r, column=c).border = box
            ws.cell(row=r, column=c).font = Font(name=FONT_NAME, size=10)
    ws.freeze_panes = "A2"
    return ws, len(df)


def kpi_card(ws, row, col, label, formula, accent, fmt=None):
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = kpi_label_font
    label_cell.alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)

    value_cell = ws.cell(row=row + 1, column=col, value=formula)
    value_cell.font = kpi_value_font
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
    if fmt:
        value_cell.number_format = fmt
    for r in range(row, row + 3):
        for c in range(col, col + 2):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor=LIGHT)
            cell.border = Border(left=Side(style="thick", color=accent),
                                  right=thin, top=thin, bottom=thin)


def main():
    wb = Workbook()
    wb.remove(wb.active)

    # ---- Data sheets (source of truth; all Dashboard formulas point here) ----
    ws_kpi, n_kpi = load_data_sheet(wb, "Data_KPI", os.path.join(ANALYSIS_DIR, "7_delivery_kpi_scorecard.csv"))
    ws_sup, n_sup = load_data_sheet(wb, "Data_SupplierDelays", os.path.join(ANALYSIS_DIR, "2_supplier_delays.csv"))
    ws_dem, n_dem = load_data_sheet(wb, "Data_HighDemand", os.path.join(ANALYSIS_DIR, "3_high_demand_products.csv"))
    ws_lead, n_lead = load_data_sheet(wb, "Data_LeadTime", os.path.join(ANALYSIS_DIR, "4_lead_time_analysis.csv"))
    ws_ful, n_ful = load_data_sheet(wb, "Data_Fulfillment", os.path.join(ANALYSIS_DIR, "5_order_fulfillment_rate.csv"))
    ws_stock, n_stock = load_data_sheet(wb, "Data_StockTrend", os.path.join(ANALYSIS_DIR, "6_stock_level_trends.csv"))
    ws_short, n_short = load_data_sheet(wb, "Data_Shortages", os.path.join(ANALYSIS_DIR, "1_inventory_shortages.csv"))
    ws_bottle, n_bottle = load_data_sheet(wb, "Data_Bottlenecks", os.path.join(ANALYSIS_DIR, "8_operational_bottlenecks.csv"))

    # ---- Dashboard sheet ----
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for col_letter in "BCDEFGHIJ":
        ws.column_dimensions[col_letter].width = 13

    ws["B2"] = "Supply Chain & Inventory Analytics"
    ws["B2"].font = title_font
    ws["B3"] = "SQL · Excel · Power BI-style reporting  |  120K+ order records across 60 suppliers, 250 SKUs, 12 warehouses"
    ws["B3"].font = subtitle_font

    # KPI cards row (row 5-7), each formula reads Data_KPI!B2 (single summary row)
    kpi_card(ws, 5, 2, "TOTAL ORDERS", "=Data_KPI!A2", TEAL, "#,##0")
    kpi_card(ws, 5, 4, "ON-TIME RATE", "=Data_KPI!B2/100", TEAL, "0.0%")
    kpi_card(ws, 5, 6, "DELAYED RATE", "=Data_KPI!C2/100", AMBER, "0.0%")
    kpi_card(ws, 5, 8, "CANCELLED RATE", "=Data_KPI!D2/100", RED, "0.0%")
    kpi_card(ws, 5, 10, "AVG DELAY (DAYS)", "=Data_KPI!F2", NAVY, "0.00")

    # Secondary KPIs computed live from the detail sheets (not just copied numbers)
    ws["B10"] = "Worst Supplier (by % Delayed)"
    ws["B10"].font = kpi_label_font
    ws["B11"] = f"=INDEX(Data_SupplierDelays!A2:A{n_sup+1}, MATCH(MAX(Data_SupplierDelays!D2:D{n_sup+1}), Data_SupplierDelays!D2:D{n_sup+1}, 0))"
    ws["B11"].font = Font(name=FONT_NAME, size=12, bold=True, color=RED)

    ws["D10"] = "Top Product by Units Ordered"
    ws["D10"].font = kpi_label_font
    ws["D11"] = f"=INDEX(Data_HighDemand!A2:A{n_dem+1}, MATCH(MAX(Data_HighDemand!D2:D{n_dem+1}), Data_HighDemand!D2:D{n_dem+1}, 0))"
    ws["D11"].font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)

    ws["F10"] = "Lowest-Fulfillment Warehouse"
    ws["F10"].font = kpi_label_font
    ws["F11"] = f"=INDEX(Data_Fulfillment!A2:A{n_ful+1}, MATCH(MIN(Data_Fulfillment!F2:F{n_ful+1}), Data_Fulfillment!F2:F{n_ful+1}, 0))"
    ws["F11"].font = Font(name=FONT_NAME, size=12, bold=True, color=AMBER)

    ws["H10"] = "Stock-outs on Latest Shortage List"
    ws["H10"].font = kpi_label_font
    ws["H11"] = f'=COUNTIF(Data_Shortages!G2:G{n_short+1}, "STOCK-OUT")'
    ws["H11"].font = Font(name=FONT_NAME, size=12, bold=True, color=RED)

    # ---- Charts ----
    # Supplier delay chart
    bar1 = BarChart()
    bar1.title = "Top 10 Suppliers by % Delayed"
    bar1.y_axis.title = "% Delayed"
    bar1.style = 10
    cats = Reference(ws_sup, min_col=1, min_row=2, max_row=11)
    data = Reference(ws_sup, min_col=4, min_row=1, max_row=11)
    bar1.add_data(data, titles_from_data=True)
    bar1.set_categories(cats)
    bar1.height, bar1.width = 8, 16
    ws.add_chart(bar1, "B14")

    # Lead time chart
    bar2 = BarChart()
    bar2.title = "Avg Lead Time: Actual vs Promised (days) by Region"
    bar2.style = 12
    cats2 = Reference(ws_lead, min_col=1, min_row=2, max_row=n_lead + 1)
    data2 = Reference(ws_lead, min_col=3, min_row=1, max_col=4, max_row=n_lead + 1)
    bar2.add_data(data2, titles_from_data=True)
    bar2.set_categories(cats2)
    bar2.height, bar2.width = 8, 16
    ws.add_chart(bar2, "G14")

    # Fulfillment by warehouse
    bar3 = BarChart()
    bar3.title = "Fulfillment Rate % by Warehouse"
    bar3.style = 11
    cats3 = Reference(ws_ful, min_col=1, min_row=2, max_row=n_ful + 1)
    data3 = Reference(ws_ful, min_col=6, min_row=1, max_row=n_ful + 1)
    bar3.add_data(data3, titles_from_data=True)
    bar3.set_categories(cats3)
    bar3.height, bar3.width = 8, 16
    ws.add_chart(bar3, "B31")

    # Stock trend line chart
    line1 = LineChart()
    line1.title = "Avg Stock Level vs Reorder Point (Monthly)"
    line1.style = 13
    cats4 = Reference(ws_stock, min_col=1, min_row=2, max_row=n_stock + 1)
    data4 = Reference(ws_stock, min_col=2, min_row=1, max_col=3, max_row=n_stock + 1)
    line1.add_data(data4, titles_from_data=True)
    line1.set_categories(cats4)
    line1.height, line1.width = 8, 16
    ws.add_chart(line1, "G31")

    ws["B48"] = "Source: sql/analysis_queries.sql -> analysis/build_summary.py -> Data_* sheets (this workbook). All Dashboard figures are live formulas, not pasted values."
    ws["B48"].font = Font(name=FONT_NAME, size=8, italic=True, color="888888")

    wb.save(OUT_PATH)
    print(f"Saved {OUT_PATH}")


if __name__ == "__main__":
    main()
